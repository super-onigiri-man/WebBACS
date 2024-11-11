from openpyxl import load_workbook
import openpyxl
import os
import sys
import PySimpleGUI as sg


import GetData
Oriconday = GetData.OriconTodays()
# コピー元のExcelファイルを開く
source_wb = load_workbook('Rank_BackUp/'+str(Oriconday)+'ベストヒットランキング.xlsx')
source_ws = source_wb.active

# コピー先のExcelファイルを開く
target_wb = load_workbook('grc.xlsx')
target_ws = target_wb['チャート']

#Noコピペ
target_ws['B2'].value = source_ws['B3'].value

#日付コピペ
target_ws['F2'].value = source_ws['F3'].value

#ThisWeekコピペ
for i in range(5, 44,2):
    source_cell = f'B{i + 1}'
    target_cell = f'B{i}'
    target_ws[target_cell].value = source_ws[source_cell].value

#LastWeekコピペ
for i in range(5, 44,2):
    source_cell = f'C{i + 1}'
    target_cell = f'C{i}'
    target_ws[target_cell].value = source_ws[source_cell].value

#Onchartコピペ
for i in range(5, 44,2):
    source_cell = f'D{i + 1}'
    target_cell = f'D{i}'
    target_ws[target_cell].value = source_ws[source_cell].value

#タイトルコピペ
for i in range(5, 44,2):
    source_cell = f'E{i + 1}'
    target_cell = f'E{i}'
    target_ws[target_cell].value = source_ws[source_cell].value

#タイトルコピペ
for i in range(5, 44,2):
    source_cell = f'F{i + 1}'
    target_cell = f'F{i}'
    target_ws[target_cell].value = source_ws[source_cell].value


def glossingover(SellNumber):

    # Bcelは今週のランキング（数字のみ）
    # Ccelは先週との比較（文字・数字混合）

    Bcel = target_ws['B'+str(SellNumber)].value
    Ccel = target_ws['C'+str(SellNumber)].value
    
    # 初登場
    if str(Ccel) == '初':
        Rank = '初登場です'

    # 再登場
    elif str(Ccel) == '再':
        Rank = '再登場です'

    # 前回と同じ
    elif Bcel == Ccel:
        Rank = '前回と同じです'

    # 前回からランクアップ
    elif Bcel < Ccel: 
        Rank = '前回'+ str(target_ws['C'+str(SellNumber)].value) + '位からアップ'

    # 前回からランクダウン
    elif Bcel > Ccel:
        Rank = '前回'+ str(target_ws['C'+str(SellNumber)].value) + '位からダウン'

    return  Rank

rank5 = glossingover(13)
rank4 = glossingover(11)
rank3 = glossingover(9)
rank2 = glossingover(7)
rank1 = glossingover(5)

# 原稿言い回し実装
target_ws = target_wb['原稿']

target_ws['C15'] = rank5
target_ws['C19'] = rank4
target_ws['C26'] = rank3
target_ws['C36'] = rank2
target_ws['C46'] = rank1

try:

    # ユーザーのダウンロードフォルダに移動
    user_folder = os.path.expanduser("~")
    folder = os.path.join(user_folder, "Downloads")
    os.chdir(folder)

    # コピー先のExcelファイルを保存
    target_wb.save(str(Oriconday)+'ベストヒット原稿.xlsx')

    # 元のディレクトリに戻る
    os.chdir(os.path.dirname(sys.argv[0]))

    # メッセージ表示
    sg.popup('原稿を作成しました',no_titlebar=True)

    # ダウンロードフォルダーを開く
    os.startfile(folder)

except Exception as e: 
    os.chdir(os.path.dirname(sys.argv[0]))
    import traceback
    with open('error.log', 'a') as f:
        traceback.print_exc( file=f)
    sg.popup_error('原稿Excelに書き込めませんでした。\n原稿Excelが開かれている可能性があります',no_titlebar=True)
