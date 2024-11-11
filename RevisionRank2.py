from openpyxl.styles import Font, PatternFill
from openpyxl import load_workbook
from openpyxl.styles.alignment import Alignment
import openpyxl
import unicodedata
import sqlite3
import csv
import PySimpleGUI as sg
import GetData


dbname = ('test2.db')
conn = sqlite3.connect(dbname, isolation_level=None)#データベースを作成、自動コミット機能ON
cursor = conn.cursor() #カーソルオブジェクトを作成


def RevisionRank(RevisionPath):
    # Excelファイルの読み込み
    excel_file = RevisionPath
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

        # データの処理と挿入
    for row in range(6, 45, 2):
        title = unicodedata.normalize("NFKC",str(sheet['E' + str(row)].value))
        artist = unicodedata.normalize("NFKC",str(sheet['F' + str(row)].value))
        rank = sheet['B' + str(row)].value
        on_chart = sheet['D'+str(row)].value
        this_number = unicodedata.normalize("NFKC",sheet['B3'].value)
        this_number = this_number.replace('No.','')
        unique_id = GetData.generate_unique_id(title,artist)

        if "再" in str(sheet['C' + str(row)].value) or "圏外" in str(sheet['C' + str(row)].value) :
            last_number = int(this_number) - 1
        elif "初" in str(sheet['C' + str(row)].value):
            on_chart = 1
            last_number = this_number
        else:
            last_number = this_number

        # データベースに挿入
        cursor.execute('''INSERT INTO music_master
        (Title, Artist, Score, Last_Rank, Last_Number, On_chart,Unique_id)
        VALUES (?, ?, 0.0, ?, ?, ?, ?)
        ON CONFLICT(Unique_id) DO UPDATE SET Last_Number = ?,Last_Rank = ?,On_Chart = ?''',
                (title, artist, rank, last_number, on_chart,unique_id,last_number,rank,on_chart))

    # コミットとクローズ
    conn.commit()

    #Scoreの値を全消し
    cursor.execute("UPDATE music_master SET Score = 0;") 

    #Last_Numbweが空文字ものを全消し
    cursor.execute("DELETE FROM music_master WHERE Last_Number = '' OR Last_Number = 0;") 

    with open('楽曲データ.csv', 'w', newline='',encoding='UTF-8') as f:
        # CSVライターオブジェクトを作成
        csv_writer = csv.writer(f)
        # ヘッダー行を書き込む
        csv_writer.writerow(['Title', 'Artist', 'Score', 'Last_Rank', 'Last_Number', 'On_Chart', 'Unique_id'])
        # データを取得
        cursor.execute("SELECT * FROM music_master")
        rows = cursor.fetchall()

        # データをCSVに書き込む
        for row in rows:
            csv_writer.writerow(row)

    # sg.popup_ok('CSVファイルに書き込みました',no_titlebar=True)